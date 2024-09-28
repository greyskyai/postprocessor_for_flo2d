import os
import re
import pandas as pd
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference

def extract_hydrograph_data(folder_path):
    """
    Extracts hydrograph data from the SWMM output file.
    
    Args:
    - file_path (str): Path to the SWMM output file.
    
    Returns:
    - dict: Dictionary of inlet data with inlet names as keys and DataFrames as values.
    """
    file_path = os.path.join(folder_path, 'SWMMQIN.OUT')
    with open(file_path, 'r') as file:
        file_content = file.readlines()
    
    inlets_data = {}
    current_inlet = None
    pattern_inlet = re.compile(r'STORM DRAIN INLET: +(.*)')
    pattern_data = re.compile(r'^\s*([\d\.]+)\s+([\d\.]+)')
    
    for line in file_content:
        inlet_match = pattern_inlet.search(line)
        data_match = pattern_data.search(line)

        if inlet_match:
            current_inlet = inlet_match.group(1).strip()
            inlets_data[current_inlet] = {'Time (hrs)': [], 'Discharge (cfs)': []}
        
        elif data_match and current_inlet:
            time = float(data_match.group(1))
            discharge = float(data_match.group(2))
            inlets_data[current_inlet]['Time (hrs)'].append(time)
            inlets_data[current_inlet]['Discharge (cfs)'].append(discharge)
    
    # Convert to DataFrame
    inlet_dfs = {inlet: pd.DataFrame(data) for inlet, data in inlets_data.items()}
    return inlet_dfs


def create_pdf_plots(inlet_dfs, output_pdf_path):
    with PdfPages(output_pdf_path) as pdf:
        sections = list(inlet_dfs.keys())
        num_pages = (len(sections) + 3) // 4

        for page in range(num_pages):
            fig, axs = plt.subplots(2, 2, figsize=(8.5, 11))
            fig.subplots_adjust(hspace=0.4, wspace=0.3)
            axs = axs.flatten()

            for i in range(4):
                idx = page * 4 + i
                if idx >= len(sections):
                    break
                inlet = sections[idx]
                df = inlet_dfs[inlet]
                
                axs[i].plot(df['Time (hrs)'], df['Discharge (cfs)'], label='Discharge', color='blue')
                axs[i].set_title(f'Inlet {inlet}')
                axs[i].set_xlabel('Time (hours)')
                axs[i].set_ylabel('Discharge (cfs)')
                axs[i].grid(True)

                peak_discharge = df['Discharge (cfs)'].max()
                time_of_peak = df['Time (hrs)'][df['Discharge (cfs)'].idxmax()]
                label = f'Peak Discharge: {peak_discharge:.2f} cfs\nTime of Peak: {time_of_peak:.2f} hrs'
                axs[i].text(0.05, 0.95, label, ha='left', va='top', transform=axs[i].transAxes, fontsize=8,
                            bbox=dict(facecolor='white', alpha=0.6))

            # Remove unused subplots
            for j in range(i + 1, 4):
                fig.delaxes(axs[j])

            pdf.savefig(fig)
            plt.close(fig)


def create_excel_with_plots(inlet_dfs, excel_path):
    """
    Creates an Excel file with hydrograph data and live plots for each inlet.
    
    Args:
    - inlet_dfs (dict): Dictionary of DataFrames containing hydrograph data.
    - excel_path (str): Path to save the Excel file.
    """
    wb = Workbook()
    wb.remove(wb.active)

    for inlet, df in inlet_dfs.items():
        # Add new sheet
        ws = wb.create_sheet(title=inlet)

        # Write DataFrame to the sheet
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Create a Line Chart
        chart = LineChart()
        chart.title = f"{inlet}\nQp = {df['Discharge (cfs)'].max():.2f} cfs, Tp = {df['Time (hrs)'][df['Discharge (cfs)'].idxmax()]:.2f} hrs"
        chart.x_axis.title = "Time (hours)"
        chart.y_axis.title = "Discharge (cfs)"
        
        # Set the data for the chart (ignoring the header row)
        data = Reference(ws, min_col=2, min_row=2, max_row=len(df) + 1)
        time = Reference(ws, min_col=1, min_row=2, max_row=len(df) + 1)
        chart.add_data(data, titles_from_data=False)
        chart.set_categories(time)

        # Add chart to sheet
        ws.add_chart(chart, "E5")

    wb.save(excel_path)

def swmm_inlet_spreadsheets_and_pdf(folder_path):
    out_folder_path = os.path.join(folder_path, 'flo2d_plots')
    pdf_output_path = os.path.join(out_folder_path, 'swmm_inlet_hydrographs.pdf')
    excel_output_path = os.path.join(out_folder_path, 'swmm_inlet_hydrographs.xlsx')

    # Step 1: Extract hydrograph data from the SWMM output file
    inlet_data = extract_hydrograph_data(folder_path)

    # Step 2: Create PDF plots with 4 plots per page
    create_pdf_plots(inlet_data, pdf_output_path)

    # Step 3: Create Excel file with hydrograph data and live plots
    create_excel_with_plots(inlet_data, excel_output_path)

if __name__ == "__main__":
    # Example file paths (update these as necessary)
    input_file = "SWMMQIN.OUT"  # Update this path
    pdf_output_path = "hydrographs_grid.pdf"
    excel_output_path = "hydrographs_data.xlsx"

    # Step 1: Extract hydrograph data from the SWMM output file
    inlet_data = extract_hydrograph_data(input_file)

    # Step 2: Create PDF plots with 4 plots per page
    create_pdf_plots(inlet_data, pdf_output_path)

    # Step 3: Create Excel file with hydrograph data and live plots
    create_excel_with_plots(inlet_data, excel_output_path)
